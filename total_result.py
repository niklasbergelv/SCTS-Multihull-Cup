#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sqlite3
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill
from openpyxl.utils import get_column_letter
from operator import itemgetter
import logging
import time

ts = time.gmtime()
timeStamp = time.strftime("%Y-%m-%d %H:%M:%S", ts)

logging.basicConfig(filename='total_result.log',level=logging.DEBUG, format='%(asctime)s %(message)s')

conn = sqlite3.connect('SCTS_SRS.db')
c = conn.cursor()

c.execute("""select race_id, race_name, race_date from races where race_date like '2018%' and race_date <= '2018-10-01' order by race_date;""")

rows = c.fetchall()

raceList = []

for row in rows:
    t = (row[0], row[1], row[2])
    race_id = row[0]
    
    #2015-10-07, only include races with participants
    c.execute("""select boat_id from results where race_id = '{}';""" . format(race_id))
    
    results = c.fetchall()
    
    if len(results) > 0: 
        raceList.append(t)
    
c.execute("""select boat_id, skipper from boats order by skipper;""")

rows = c.fetchall()

boatList = []

for row in rows:
    races = dict()
    t = (row[0], row[1], races)
    boatList.append(t)

for iterator in range(len(boatList)):
    boat_id, skipper, races = boatList[iterator]
    totalScore = 0
    
    for race in raceList:
        d = dict()
        race_id, race_name, race_date = race
        c.execute("""select position, points from results where race_id = '{}' and boat_id = '{}';""" . format(race_id, boat_id))
        row = c.fetchone()
        #print("race_name:{}, skipper:{}, position:{}\n" . format(race_name, skipper, row[0]))
        if row != None:
            print("race_name:{}, skipper:{}, position:{}\n" . format(race_name, skipper, row[0]))
            d['position'] = row[0]
            d['points'] = row[1]
            if row[1] != None:
                totalScore += row[1]
        else:
            d['position'], d['points'] = ('-', '-')
        races[race_name] = d
    
    boatList[iterator] = (boat_id, skipper, totalScore, races)
    
boatList.sort(key=lambda tup:tup[2], reverse=False)
        
print(json.dumps(boatList, indent=4))

#create excel output

wb = Workbook()

membershipDict = dict()

certificatesDict = dict()


for iterator in range(len(raceList)):
    result = []
    result2 = []
    race = raceList[iterator]
    race_id = race[0]
    race_name = race[1]
    race_date = race[2]
    
    print("race_name: '{}', race_date: '{}'" .format(race_name, race_date))
    
    for iter2 in range(len(boatList)):
        boat = boatList[iter2]
        boat_id = boat[0]
        skipper = boat[1]
        
        
        boat_name = ""
        membership_no = ""
        certificates_id = ""
        
        c.execute("""select boat_name from boats where boat_id = '{}';""" . format(boat_id))
        row = c.fetchone()
        if row != None:
            boat_name = row[0]
        
        #2017-10-16 change so that only members that has paid to a certain yer, (use join?)
        #c.execute("""select medlemsnummer from matrikel where boat_id = '{}';""" . format(boat_id))
        #c.execute("""select membershipNumber from membership where boat_id = '{}';""" . format(boat_id))
        c.execute("""select membershipNumber from members_2018 where boat_id = '{}';""" . format(boat_id))
        row = c.fetchone()
        if row != None:
            membership_no = row[0]
            membershipDict[boat_id] = membership_no
        
        #2018-10-10, select cert id valid for race date
        #2018-10-07, pick certid
        #2016-10-16, add certificiate id
        c.execute("""select certid, valid_from, valid_to, SRS, shorthanded, srs_without, shorthanded_without from certificates where boat_id = '{0}' and valid_from <= '{1}' and valid_to >= '{1}' ;""" . format(boat_id, race_date))
        #c.execute("""select certid, valid_from, valid_to from certificates where boat_id = '{0}' and valid_from <= '{1}' and valid_to >= '{1}' ;""" . format(boat_id, race_date))
        #c.execute("""select certificates_id, valid_from, valid_to from certificates where boat_id = '{0}' and valid_from <= '{1}' and valid_to >= '{1}' ;""" . format(boat_id, race_date))
        row = c.fetchone()
        if row != None:
            certificates_id = row[0]
            valid_from = row[1]
            valid_to = row[2]
            SRS = row[3]
            shorthanded = row[4]
            srs_without = row[5]
            shorthanded_without = row[6]
            certificatesDict[boat_id] = row
            #print("race_name: '{}', boat_id: '{}', certificiates_id: '{}', race_date: '{}', valid_from: '{}', valid_to: '{}'" .format(race_name, boat_id, certificiates_id, race_date, valid_from, valid_to))
        else:
            print("No valid certificiate; race_name: '{}', boat_id: '{}'" . format(race_name, boat_id))
        
        
        c.execute("""select SRS, sailed_time, counted_time, position, points from results where race_id = '{}' and boat_id = '{}';""" . format(race_id, boat_id))
        row = c.fetchone()
        if row != None:
            (SRS, sailed_time, counted_time, position, points) = row
            if type(position) == int:
                
                if boat_id not in certificatesDict:
                    result.append([boat_name, skipper, membership_no, "", SRS, sailed_time, counted_time, position, points])
                else:
                    certid_info = certificatesDict[boat_id]
                    result.append([boat_name, skipper, membership_no, certid_info[0], SRS, sailed_time, counted_time, position, points, certid_info[3], certid_info[4], certid_info[5], certid_info[6]])
            else:
                if sailed_time != 'DNS':
                    if boat_id not in certificatesDict:
                        #2015-10-07 only include result if boat has started
                        result2.append([boat_name, skipper, membership_no, "", SRS, sailed_time, counted_time, position, points])
                    else:
                        certid_info = certificatesDict[boat_id]
                        result2.append([boat_name, skipper, membership_no, certificatesDict[boat_id][0], SRS, sailed_time, counted_time, position, points, certid_info[3], certid_info[4], certid_info[5], certid_info[6]])
    
    
    result.sort(key = itemgetter(7, 1))
    ws = wb.create_sheet()
    ws.title = ("{}" . format(race_name))
    ws.append(["båtnamn", "skeppare", "SCTS no", "Cert id", "Anmält SRS", "Seglad tid", "Beräknad tid", "Placering", "Poäng", "SRS", "SH", "SRS utan", "SH utan"])
    rowNum = 2
    for iterator in result: 
        ws.append(iterator)
        
        #if boat_id not in certificatesDict:
        if iterator[3] == "":
            logging.info('no cert id given for boat {0} in race {1}' . format(iterator[0], race_name))
            for colNum in range(1,14):
                ws.cell(row = rowNum, column=colNum).fill = PatternFill(fgColor = "ff0000", fill_type = 'solid')
                
        else:
            certificate_SRS = []
            certificate_SRS.append(iterator[9])
            certificate_SRS.append(iterator[10])
            certificate_SRS.append(iterator[11])
            certificate_SRS.append(iterator[12])
            
            if iterator[4] not in certificate_SRS:
                logging.info('Anmält SRS, {0} do not match cert {1} for {2} in {3}' . format(iterator[4], iterator[3],iterator[0], race_name))
                for colNum in range(1,14):
                    ws.cell(row = rowNum, column=colNum).fill = PatternFill(fgColor = "ff0000", fill_type = 'solid')
                
        rowNum += 1
        
    if len(result2) > 0:
        result2.sort(key = itemgetter(1))
        for iterator in result2:
            ws.append(iterator)
            
pointsAccumulated = []

columns = ["skeppare", "båtnamn", "SCTS no"]
#columns2 = ["skeppare", "båtnamn", "SCTS no"]
#columns3 = ["skeppare", "båtnamn", "SCTS no"]


for race in raceList:
    (race_id, race_name, race_date) = race
    columns.append(race_name)
    #columns2.append(race_name)
    #columns3.append(race_name)

#c.execute("""select boat_id, skipper, boat_name from boats order by skipper;""")
c.execute("""select boat_id, skipper, boat_name from boats where boat_id IN (select boat_id from results where race_id IN (select race_id from '{}'));""" . format('races_2018'))

boatResults = []
boatPositions = []
boatPoints = []

rows = c.fetchall()

for row in rows:
    (boat_id, skipper, boat_name) = row
    
    membership_no = ""
    
    
    #if (boat_id in membershipDict) and (boat_id in certificatesDict): boat where the skipper is a member should be part of the overall result
    if (boat_id in membershipDict):
        membership_no = membershipDict[boat_id]
        
        
        
        #2015-10-07, only add boat to accumulated list
        result = [skipper, boat_name, membership_no]
        result2 = [skipper, boat_name, membership_no]
        result3 = [skipper, boat_name, membership_no]
        pointsAcc = 0
        for race in raceList: 
            (race_id, race_name, race_date) = race
            c.execute("""select points, position, sailed_time from results where race_id = {} and boat_id = {};""" . format(race_id, boat_id))
            row = c.fetchone()
            points = '-'
            position = '-'
            if row != None:
                points = row[0]
                position = row[1]
                sailed_time = row[2]
                
                if points != None:
                    pointsAcc += points
                if sailed_time == 'DNS':
                    points = '-' 
            result.append(pointsAcc)
            result2.append(position)
            result3.append(points)
         
        boatResults.append(result)
        boatPositions.append(result2)
        boatPoints.append(result3)
  
ws = wb.create_sheet()
ws.title = ("accumulerade poäng")

columns2 = ["position"] + columns

ws.append(columns2)
#ws.append(columns)


#sort boatResults

boatResults.sort(key=lambda tup: tup[16], reverse=True)
    
total_Position = 1

ackPoints = 0

i = 1

prev = 0

for result in boatResults:
    
    if i  == 1:
        prev = i
        ws.append([i] + result)
    else:
        if  ackPoints == result[16]:
            ws.append([prev] + result)
        else:
            prev = i
            ws.append([i] + result)
    
    ackPoints = result[16]
   
    i += 1
    
    
# ====================================
    
ws = wb.create_sheet()
ws.title = ("placeringar")
ws.append(columns)

for result in boatPositions:
    ws.append(result)
    
ws = wb.create_sheet()
ws.title = ("poäng")

columns2 = []
columns2.append("Plac.") 
columns2.extend(columns) 
columns2.append("Totalt")

ws.append(columns2)

#ws.append(columns)

logging.info("Columns in sheet: Poäng: " . format(columns))

i = 2
for result in boatPoints:
    #temp = "=SUMMA(D{0}:S{0})" . format(i)
    #temp = "plats för formel"
    #result.append(temp)
    temp = []
    temp.append("")
    temp.extend(result)
    #ws.append(result)
    ws.append(temp)
    
#ws["T2"] = "=SUMMA(D2:S2)"



ws = wb.create_sheet()
ws.title = ("Certificates")
c.execute("""PRAGMA table_info(certificates);""")
rows = c.fetchall()

certificates_columns = []

for row in rows:
    (cid, col_name, col_type, not_null, dflt_value, pk) = row
    certificates_columns.append(col_name)


temp = certificates_columns[0]
certificates_columns[0] = certificates_columns[11]
certificates_columns[11] = temp
    
ws.append(certificates_columns)

c.execute("""select * from certificates where valid_from like \"2018%\";""")
rows = c.fetchall()

for row in rows:
    #(certificates_id, boat_id, owner, sail_number, boat_name, boat_type, SRS, SRS2, valid_from, valid_to, creation_date, certid, shorthanded, srs_without, shorthanded_without) = split(row)
    column_temp = []
    
    for col in row:
        column_temp.append(col)
    
    temp = column_temp[0]   
    column_temp[0] = column_temp[11]
    column_temp[11] = temp
    ws.append(column_temp)
    
    
# adjust column sizes

logging.info("Sheet names:")

nr = 1

for sheet in wb.sheetnames:
    logging.info("{} {} number of columns: {}, rows: {}" . format(nr, sheet, str(wb[sheet].max_column), str(wb[sheet].max_row)))
    nr+=1
    worksheet = wb[sheet]
    column_widths = []
    for row in worksheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(cell.value))
    logging.info("Column widths {}" . format(str(column_widths)))
    
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width
   
        
wb.save("{}{}{}" . format("SCTS_RACES_", timeStamp, ".xlsx"))              
#wb.save("SCTS_RACES_2018-12-28.xlsx")        
            
    
conn.close()